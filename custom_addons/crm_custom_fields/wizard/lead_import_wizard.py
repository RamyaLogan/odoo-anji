from odoo import models, fields, api,SUPERUSER_ID
from odoo.exceptions import UserError
from openpyxl import load_workbook
import base64
import os
import logging
from datetime import datetime, timedelta
import openpyxl
from dateutil.relativedelta import relativedelta
from psycopg2.extras import execute_values
from odoo.api import Environment
import psycopg2
from odoo.modules.registry import Registry
import boto3
import io
from collections import defaultdict, deque

S3_BUCKET = 'mhs-doneztech'
S3_PREFIX = 'crm-imports/' 
_logger = logging.getLogger(__name__)

class LeadImportWizard(models.TransientModel):
    _name = 'lead.import.wizard'
    _description = 'Lead Import Wizard'

    file = fields.Binary('Excel File', required=True)
    filename = fields.Char('File Name')
    saved_filename = fields.Char('Saved File Name', readonly=True)
    lead_source = fields.Selection([
        ('web', 'Website'),
        ('fb', ' Facebook'),
    ], string='Lead Source', required=True, default='web')
    import_type = fields.Selection([('lead', 'Lead'), ('opportunity', 'Opportunity')],default='lead',
        string='Import Type',
        invisible=True,)
        # Hot lead split
    split_days = fields.Integer(string="Split Over Days", required=True, default=1)
    percent_hot_senior = fields.Integer(string="Hot % - Senior", default=60)
    percent_hot_junior = fields.Integer(string="Hot % - Junior", default=30)
    percent_hot_trainee = fields.Integer(string="Hot % - Trainee", default=10)

    # Warm lead split
    percent_warm_senior = fields.Integer(string="Warm % - Senior", default=20)
    percent_warm_junior = fields.Integer(string="Warm % - Junior", default=40)
    percent_warm_trainee = fields.Integer(string="Warm % - Trainee", default=40)

    @api.constrains('percent_hot_senior', 'percent_hot_junior', 'percent_hot_trainee',
                    'percent_warm_senior', 'percent_warm_junior', 'percent_warm_trainee')
    def _check_percentage_total(self):
        for rec in self:
            hot_total = rec.percent_hot_senior + rec.percent_hot_junior + rec.percent_hot_trainee
            warm_total = rec.percent_warm_senior + rec.percent_warm_junior + rec.percent_warm_trainee
            if hot_total != 100:
                raise UserError("Hot lead distribution must total 100%.")
            if warm_total != 100:
                raise UserError("Warm lead distribution must total 100%.")
    @api.onchange('file')
    def _onchange_file(self):
        if self.filename:
            base, ext = os.path.splitext(self.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.saved_filename = f"{base}_{timestamp}{ext}"


    def action_import_leads(self):
        s3 = boto3.client('s3')
        for record in self:
            if record.file and record.filename:
                base, ext = os.path.splitext(record.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                s3_key = f"{S3_PREFIX}{base}_{timestamp}{ext}"

                decoded_file = base64.b64decode(record.file)
                s3.upload_fileobj(io.BytesIO(decoded_file), S3_BUCKET, s3_key)

                # Trigger the job with the S3 path
                self.env['lead.import.wizard'].with_delay(description="CRM S3 Import").process_uploaded_leads_from_s3(s3_key, self.import_type, self.lead_source,self.split_days, self.percent_hot_senior, self.percent_hot_junior, self.percent_hot_trainee,
                                                                                                                self.percent_warm_senior, self.percent_warm_junior, self.percent_warm_trainee)
    
    @api.model
    def process_uploaded_leads_from_s3(self, s3_key, import_type, lead_source,split_days,percent_hot_senior=60, percent_hot_junior=30, percent_hot_trainee=10,
                                       percent_warm_senior=20, percent_warm_junior=40, percent_warm_trainee=40):
        import tempfile
        s3 = boto3.client('s3')
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            s3.download_fileobj(S3_BUCKET, s3_key, tmp)
            tmp.flush()
            os.fsync(tmp.fileno())  # Ensure data is fully written to disk
            tmp_path = tmp.name

        try:
            return self.process_uploaded_leads(tmp_path, import_type,lead_source) if import_type == 'lead' else self.process_uploaded_opportunity(tmp_path,percent_hot_senior, percent_hot_junior, percent_hot_trainee,
                                     percent_warm_senior, percent_warm_junior, percent_warm_trainee,split_days)
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    
    def process_uploaded_leads(self, file_path, import_type, lead_source):
        team_name = 'Online Sales Team - Web' if lead_source == 'web' else 'Online Sales Team - FB'
        online_team = self.env['crm.team'].search([('name', '=', team_name)], limit=1)
        assigned_users = online_team.member_ids.ids
        user_count = len(assigned_users)

        header_map, data_rows = self._load_excel(import_type, file_path)
        phones_to_import = self._extract_phones(data_rows, header_map)
        existing_leads = self.env['crm.lead'].search([('phone', 'in', list(phones_to_import))])
        existing_lead_map = {self.normalize_phone(lead.phone): lead for lead in existing_leads}

        leads_to_create, activities_to_create = [], []
        duplicate_records, imported, duplicates, index = set(), 0, 0, 0
        first_batch_code = data_rows[0][header_map.get('batch_code')] if data_rows else None
        if first_batch_code:
            self.env['lead.import.batch'].create({
                'name': first_batch_code,
                'import_type': 'lead',
            })
        for row in data_rows:
            phone = self.normalize_phone(row[header_map['phone']])
            if not phone:
                continue
            if phone in duplicate_records:
                duplicates += 1
                continue

            name = row[header_map['name']]
            email = row[header_map['email']]
            whatsapp_no = row[header_map['whatsapp_no']]
            batch_code = row[header_map.get('batch_code')]
            sugar_level = row[header_map.get('sugar_level')]
            call_status = 'new'

            lead = existing_lead_map.get(phone)
            if lead:
                lead.write({
                    'name': name,
                    'email_from': email,
                    'whatsapp_no': whatsapp_no,
                    'batch_code_full': batch_code,
                    'sugar_level': sugar_level,
                    'call_status': call_status
                })
                continue

            assigned_user_id = assigned_users[index % user_count] if user_count > 0 else False

            leads_to_create.append({
                'name': name,
                'phone': phone,
                'email_from': email,
                'user_id': assigned_user_id,
                'type': 'lead',
                'whatsapp_no': whatsapp_no,
                'batch_code_full': batch_code,
                'sugar_level': sugar_level,
                'call_status': call_status
            })
            duplicate_records.add(phone)
            index += 1

            if len(leads_to_create) >= 500:
                imported += self._create_leads_and_activities(leads_to_create)
                leads_to_create, activities_to_create = [], []

        self.create_followup_activities(existing_leads, summary="Second Time - Recall", days=1)

        if leads_to_create:
            imported += self._create_leads_and_activities(leads_to_create)

        os.remove(file_path)
        _logger.info(f"Leads imported: {imported}, Duplicates: {duplicates}, Existing Leads: {len(existing_leads)}")
        return {
            'imported': imported,
            'existing_leads': len(existing_leads),
            'duplicates': duplicates
        }

    def process_uploaded_opportunity(self, file_path,percent_hot_senior, percent_hot_junior, percent_hot_trainee,
                                     percent_warm_senior, percent_warm_junior, percent_warm_trainee,split_days):
        self = self.with_env(self.env(user=SUPERUSER_ID))
        team = self.env['crm.team'].search([('name', '=', 'Offline Sales Team')], limit=1)
        if not team:
            raise UserError("Offline Sales Team not found.")

        users_by_role = defaultdict(list)
        for user in team.member_ids:
            if user.role_level:
                users_by_role[user.role_level].append(user.id)

        role_weights = {
            'hot': {
                'senior': percent_hot_senior,
                'junior': percent_hot_junior,
                'trainee': percent_hot_trainee
            },
            'warm': {
                'senior': percent_warm_senior,
                'junior': percent_warm_junior,
                'trainee': percent_warm_trainee
            }
        }

        header_map, data_rows = self._load_excel('opportunity', file_path)
        phones = list(self._extract_phones(data_rows, header_map))
        existing = self.env['crm.lead'].search([('phone', 'in', phones)])
        existing_map = {self.normalize_phone(l.phone): l for l in existing}

        tag_cache = {}
        for tag in ['hot', 'warm', 'prospect', 'client']:
            t = self.env['crm.tag'].search([('name', '=', tag)], limit=1)
            if not t:
                t = self.env['crm.tag'].create({'name': tag})
            tag_cache[tag] = t

        assignments_by_day = defaultdict(list)
        category_buckets = {'hot': [], 'warm': [],'prospect': [], 'client': []}
        seen_phones = set()

        updated, created, skipped = 0, 0, 0
        first_batch_code = data_rows[0][header_map.get('batch_code')] if data_rows else None
        if first_batch_code:
            self.env['lead.import.batch'].create({
                'name': first_batch_code,
                'import_type': 'opportunity',
            })
        user_map = {}
        for row in data_rows:
            try:
                phone = self.normalize_phone(row[header_map['phone']])
                if not phone or phone in seen_phones:
                    skipped += 1
                    continue

                seen_phones.add(phone)
                category = row[header_map['category']].strip().lower()
                salesperson_name = row[header_map['salesperson_name']]
                user_id = None

                if salesperson_name:
                    if salesperson_name in user_map:
                        user_id = user_map[salesperson_name]
                    else:
                        user = self.env['res.users'].search([('name', '=', salesperson_name)], limit=1)
                        user_id = user.id if user else False
                        user_map[salesperson_name] = user_id  # Store even if False to avoid rechecking

                lead = existing_map.get(phone)
                if not lead:
                    lead = self.env['crm.lead'].with_context(skip_remarks_check=True,skip_followup_creation=True).create({
                        'name': row[header_map['name']],
                        'phone': phone,
                        'type': 'opportunity',
                        'sugar_level': row[header_map['sugar_level']],
                        'status': row[header_map['stage']],
                        'batch_code_full': row[header_map['batch_code']],
                        'email_from': row[header_map.get('email')],
                        'call_status': 'new',
                        'user_id': user_id
                    })
                    created += 1
                elif lead.type == 'opportunity' :
                    lead.with_context(skip_remarks_check=True,skip_followup_creation=True).write({
                        'name': row[header_map['name']],
                        'sugar_level': row[header_map['sugar_level']],
                        'status': 'new',
                        'batch_code_full': row[header_map['batch_code']],
                        'email_from': row[header_map.get('email')]
                    })
                    if lead.user_id:
                        self.env['mail.activity'].create({
                            'res_model_id': self.env['ir.model']._get_id('crm.lead'),
                            'res_id': lead.id,
                            'activity_type_id': self.env.ref('mail.mail_activity_data_call').id,
                            'summary': "Second Time - Recall",
                            'user_id': lead.user_id.id,
                            'date_deadline': fields.Date.today() + timedelta(days=2),
                        })
                        updated += 1
                        continue
                else:
                    lead.with_context(skip_remarks_check=True,skip_followup_creation=True).write({
                        'name': row[header_map['name']],
                        'type': 'opportunity',
                        'sugar_level': row[header_map['sugar_level']],
                        'status': row[header_map['stage']],
                        'batch_code_full': row[header_map['batch_code']],
                        'email_from': row[header_map.get('email')],
                        'call_status': 'new',
                        'user_id': user_id
                    })
                    updated += 1

                lead.with_context(skip_remarks_check=True,skip_followup_creation=True).write({'tag_ids': [(4, tag_cache[category].id)]})

                category_buckets[category].append({
                    'id': lead.id,
                    'phone': phone,
                    'category_tag': tag_cache[category].id,
                    'day': None
                })

            except Exception as e:
                skipped += 1
                _logger.warning(f"Skipping row due to error: {e}")

        def assign_opportunities(tag, role_weights):
            pool = category_buckets[tag]
            total = len(pool)
            if total == 0:
                return

            fractions = {}
            counts = {}
            for role in ['senior', 'junior', 'trainee']:
                exact = (role_weights[role] / 100) * total
                counts[role] = int(exact)
                fractions[role] = exact - counts[role]

            assigned = sum(counts.values())
            leftover = total - assigned
            leftover_roles = sorted(['senior', 'junior', 'trainee'], key=lambda r: (-fractions[r], -role_weights[r]))
            for i in range(leftover):
                counts[leftover_roles[i % len(leftover_roles)]] += 1

            cursor = 0
            for role in ['senior', 'junior', 'trainee']:
                users = users_by_role[role]
                if not users:
                    continue
                user_queue = deque(users)

                per_day = counts[role] // split_days
                leftover_day = counts[role] % split_days

                for day in range(split_days):
                    daily_count = per_day + (1 if day < leftover_day else 0)
                    for _ in range(daily_count):
                        if cursor >= total:
                            break
                        record = pool[cursor]
                        record['assignee_id'] = user_queue[0]
                        record['day'] = day
                        assignments_by_day[day].append(record)
                        user_queue.rotate(-1)
                        cursor += 1
        assign_opportunities('hot', role_weights['hot'])
        assign_opportunities('warm', role_weights['warm'])
            
        for day, assignments in assignments_by_day.items():
            eta = fields.Datetime.now() + timedelta(days=day)
            _logger.info(f"Queuing job for Day {day + 1} with {len(assignments)} assignments. ETA: {eta}")
            self.env['lead.import.wizard'].with_delay(eta=eta).finalize_opportunity_assignment(assignments)

        os.remove(file_path)
        _logger.info(f"Conversion complete. Updated: {updated}, Created: {created}, Skipped: {skipped}, Existing: {len(existing_map)}")
        return {
            'updated': updated,
            'created': created,
            'skipped': skipped,
            'existing_leads': len(existing_map),
        }

    @api.model
    def finalize_opportunity_assignment(self, assignment_list):
        for a in assignment_list:
            lead = self.env['crm.lead'].browse(a['id'])
            if not lead:
                continue
            lead.write({
                'user_id': a['assignee_id']
            })
            lead.write({'tag_ids': [(4, a['category_tag'])]})
            self.env['mail.activity'].create({
                'res_model_id': self.env['ir.model']._get_id('crm.lead'),
                'res_id': lead.id,
                'activity_type_id': self.env.ref('mail.mail_activity_data_call').id,
                'summary': "Post Conversion Follow-up",
                'user_id': lead.user_id.id,
                'date_deadline': fields.Date.today() + timedelta(days=2),
            })

    def _load_excel(self,import_type, file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_map = {header.lower().strip(): idx for idx, header in enumerate(header_row)}

        required_cols = ['name', 'phone', 'email', 'whatsapp_no','sugar_level','batch_code'] if import_type == 'lead' else ['name', 'phone', 'category', 'sugar_level', 'stage', 'batch_code']
        for col in required_cols:
            if col not in header_map:
                raise UserError(f"Missing required column: {col}")

        data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        return header_map, data_rows

    def _extract_phones(self, data_rows, header_map):
        phones = set()
        for row in data_rows:
            try:
                phone = row[header_map['phone']]
                if phone:
                    phones.add(self.normalize_phone(phone))
            except IndexError:
                continue
        return phones

    def _safe_int(self, val):
        try:
            return int(val)
        except:
            return 0

    def _create_leads_and_activities(self, leads_data):
        created = self.env['crm.lead'].create(leads_data)
        self.create_followup_activities(created, summary="Follow Up Call", days=1)
        return len(created)

    def create_followup_activities(self, leads, summary='Second Time - Recall', days=1):
        Activity = self.env['mail.activity'].sudo().with_context(mail_notrack=True, tracking_disable=True, mail_create_nolog=True)
        crm_model_id = self.env['ir.model']._get_id('crm.lead')
        activity_type = self.env.ref('mail.mail_activity_data_call')
        deadline = fields.Date.context_today(self) + relativedelta(days=days)

        activities_to_create = []
        for lead in leads:
            activities_to_create.append({
                'res_model_id': crm_model_id,
                'res_id': lead.id,
                'activity_type_id': activity_type.id,
                'summary': summary,
                'user_id': lead.user_id.id,
                'date_deadline': deadline,
            })
        if len(activities_to_create) >= 500:
                Activity.create(activities_to_create)
                activities_to_create = []
        if activities_to_create:
            Activity.create(activities_to_create)

    def _prepare_activity(self, lead, user_id,summary, activity_type_id):
        return {
            'res_model_id': self.env['ir.model']._get_id('crm.lead'),
            'res_id': lead.id,
            'activity_type_id': activity_type_id,
            'summary': summary,
            'user_id': user_id,
            'date_deadline': fields.Date.context_today(self) + relativedelta(days=1),
        }

    def normalize_phone(self,phone):
        return ''.join(filter(str.isdigit, str(phone).strip()))

    def is_valid_phone(self, phone):
        return True
        # try:
        #     parsed = phonenumbers.parse(phone, None)
        #     return phonenumbers.is_valid_number(parsed)
        # except:
        #     return False